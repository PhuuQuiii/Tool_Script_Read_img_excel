#!/usr/bin/env python3
"""
Fill Excel from image using Claude Vision API.
Usage: python3 fill_excel.py <image_path> [excel_path]
"""

import sys
import os
import base64
import json
import re
from datetime import datetime
import openpyxl
import anthropic

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "Fill hist 2.xlsx")
SHEET_NAME = "Sheet1"

PROMPT = """
Look at this image carefully. Find and extract these 4 fields:
1. PO No. (Purchase Order Number) - usually looks like BJQ... or similar alphanumeric code
2. GR Date (Goods Receipt Date) - a date in format dd/mm/yyyy
3. WorkScope / Scope - short code like REP, RCT, OHC, RBTH, etc.
4. Vendor Name - company name

Return ONLY a JSON object with exactly these keys (no extra text):
{
  "po_no": "...",
  "gr_date": "...",
  "scope": "...",
  "vendor_name": "..."
}

If a field cannot be found, use null.
The date should be in DD/MM/YYYY format.
"""


def encode_image(image_path: str) -> tuple[str, str]:
    ext = os.path.splitext(image_path)[1].lower()
    media_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
                 ".gif": "image/gif", ".webp": "image/webp"}
    media_type = media_map.get(ext, "image/png")
    with open(image_path, "rb") as f:
        return base64.standard_b64encode(f.read()).decode("utf-8"), media_type


def extract_data_from_image(image_path: str) -> dict:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY not set. Run: export ANTHROPIC_API_KEY=your_key")

    client = anthropic.Anthropic(api_key=api_key)
    image_data, media_type = encode_image(image_path)

    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=512,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_data}},
                {"type": "text", "text": PROMPT}
            ]
        }]
    )

    text = message.content[0].text.strip()
    # Extract JSON even if wrapped in markdown code block
    match = re.search(r'\{.*\}', text, re.DOTALL)
    if match:
        return json.loads(match.group())
    raise ValueError(f"Could not parse JSON from response: {text}")


def parse_date(date_str: str):
    if not date_str:
        return None
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            continue
    return None


def fill_excel(data: dict, excel_path: str = EXCEL_FILE):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[SHEET_NAME]

    # Find next empty data row (row 1 is header, data starts at row 2)
    next_row = 2
    for row in ws.iter_rows(min_row=2, max_col=4):
        if any(cell.value for cell in row):
            next_row = row[0].row + 1
        else:
            break

    row_num = next_row - 1  # sequential number for column B

    # Column mapping (1-indexed): B=2, C=3, D=4, E=5, F=6, G=7, H=8, I=9, J=10, K=11
    ws.cell(row=next_row, column=2, value=row_num)
    ws.cell(row=next_row, column=3, value=") ")
    ws.cell(row=next_row, column=4, value=data.get("po_no"))
    ws.cell(row=next_row, column=5, value=" GRD ")
    gr_date = parse_date(data.get("gr_date"))
    ws.cell(row=next_row, column=6, value=gr_date)
    ws.cell(row=next_row, column=7, value=" (")
    ws.cell(row=next_row, column=8, value=data.get("scope"))
    ws.cell(row=next_row, column=9, value=" - ")
    ws.cell(row=next_row, column=10, value=data.get("vendor_name"))
    ws.cell(row=next_row, column=11, value=")")
    # TEXTJOIN formula in column A
    ws.cell(row=next_row, column=1,
            value=f'=_xlfn.TEXTJOIN("",1,B{next_row}:E{next_row},TEXT(F{next_row},"dd/mm/yyyy"),G{next_row}:K{next_row})')

    wb.save(excel_path)
    print(f"✓ Added row {row_num} to {excel_path}")
    print(f"  PO No:       {data.get('po_no')}")
    print(f"  GR Date:     {data.get('gr_date')} → {gr_date.strftime('%d/%m/%Y') if gr_date else 'INVALID'}")
    print(f"  Scope:       {data.get('scope')}")
    print(f"  Vendor Name: {data.get('vendor_name')}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 fill_excel.py <image_path> [excel_path]")
        sys.exit(1)

    image_path = sys.argv[1]
    excel_path = sys.argv[2] if len(sys.argv) > 2 else EXCEL_FILE

    if not os.path.exists(image_path):
        print(f"Error: Image not found: {image_path}")
        sys.exit(1)

    print(f"Reading image: {image_path}")
    data = extract_data_from_image(image_path)
    print(f"Extracted: {json.dumps(data, indent=2)}")
    fill_excel(data, excel_path)


if __name__ == "__main__":
    main()
